"""Importação da planilha de matrículas da escola ("Lista Piloto").

Reproduz em memória a estrutura real: uma aba por turma com cabeçalho
(Escola / Lista Piloto / Professor / Ano / Turno / Nº da Classe) e, a
partir da linha de títulos, uma linha por aluno com muitos campos. Uma
aba de resumo ("QE") sem coluna "Nome" deve ser ignorada.
"""
import io
from datetime import date

import openpyxl
from sqlalchemy import func, select

from app.models import (Aluno, Importacao, LogAuditoria, Matricula,
                        SnapshotMatific, Turma)
from app.services.importacao import casa_abreviado_posicional, tokens_nome

CT_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
COLUNAS = ["N.º", "RM", "RA", "Nome", "Data de Nasc.", "Responsável",
           "Endereço", "Bairro", "Telefone", "RG", "CPF", "SUS",
           "SEXO", "RAÇA COR", "BOLSA FAMÍLIA"]


def _aba(ws, turma, turno, professor, sed, alunos):
    ws.append([None, None, None, "PREFEITURA MUNICIPAL DE CARAGUATATUBA"])
    ws.append(["Escola:", None, "EMEF PROFESSOR JORGE PASSOS", None, None,
               None, None, "Lista Piloto:", 2026])
    ws.append([])
    ws.append(["Professor:", None, professor, None, "Ano:", turma,
               f"Turno:       {turno}", None, "N.º da Classe (SED):", sed])
    ws.append([])
    ws.append(COLUNAS)
    for linha in alunos:
        ws.append(linha)


def _planilha() -> bytes:
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "1º Ano A"
    _aba(a, "1º Ano A", "TARDE", "PAULA", 300396614, [
        [1, 4446, "123.100.026-0", "AGATHA VITORIA MOURA DA SILVA",
         date(2019, 9, 18), "BRENDA", "RUA X, 178", "JARAGUAZINHO",
         "12-98109-4185", "66.4-1", "574.900.168-78", "700 1", "F", "PARDA", "SIM"],
        [2, 4447, "121.721.835-X", "ALICE DE ALMEIDA LEAL",
         date(2019, 5, 2), "MICHELLE", "RUA Y, 918", "JARAGUAZINHO",
         "99753-5290", "", "", "", "F", "BRANCA", "NÃO"],
    ])
    b = wb.create_sheet("1º Ano B")
    _aba(b, "1º ANO B", "TARDE", "MAIARA", 300396635, [
        [1, 4468, "121.521.972-6", "ADRYAN RIBEIRO VALERIANO",
         date(2018, 3, 10), "PRISCILA", "RUA Z, 453", "RIO DO OURO",
         "12 99608-2283", "65.8-X", "569.950.318-81", "700 0", "M", "BRANCA", "SIM"],
    ])
    qe = wb.create_sheet("QE")  # resumo — sem coluna "Nome" → ignorado
    qe.append(["Nr. Classe", "Matriculados", "Transferidos", "Ativos"])
    qe.append(["1º ano A", 300396614, 2, 2])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _upload(nome="Lista Piloto 2026.xlsx"):
    return {"arquivo": (nome, _planilha(), CT_XLSX)}


def test_parser_reconhece_turmas_alunos_e_ficha():
    from app.services.lista_piloto import analisar_matriculas
    a = analisar_matriculas(_planilha(), "Lista Piloto 2026.xlsx")
    assert a.escola_detectada.startswith("EMEF PROFESSOR JORGE PASSOS")
    assert a.ano_letivo == 2026
    assert [t.nome for t in a.turmas] == ["1º Ano A", "1º Ano B"]   # QE ignorada
    t = a.turmas[0]
    assert t.ano_escolar == "1º Ano" and t.turno == "tarde"
    assert t.professor == "PAULA" and t.sed == "300396614"
    al = t.alunos[0]
    assert al.nome == "AGATHA VITORIA MOURA DA SILVA"
    assert al.numero_chamada == 1 and al.data_nascimento == "2019-09-18"
    # Minimização (LGPD): só o RA é guardado; sensíveis/documentos são ignorados.
    assert al.ficha == {"ra": "123.100.026-0"}
    assert "cpf" not in al.ficha and "raca_cor" not in al.ficha
    assert "sus" not in al.ficha and "responsavel" not in al.ficha


def test_analisar_marca_turmas_existentes(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    r = cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/analisar",
                     files=_upload())
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["total_turmas"] == 2 and corpo["total_alunos"] == 3
    nomes = {t["nome"]: t for t in corpo["turmas"]}
    # A fixture já tem "3º Ano A"; as da planilha são novas.
    assert nomes["1º Ano A"]["ja_existe"] is False
    assert nomes["1º Ano A"]["exemplos"][0] == "AGATHA VITORIA MOURA DA SILVA"


def test_previa_resume_novos_atualizar_e_duplicidade(cliente, db, escola_completa):
    """Etapa 1: o resumo antes da confirmação separa novos × a atualizar e marca
    duplicidade de turma; reimportar reconhece os já cadastrados (atualizar)."""
    escola_id = escola_completa["escola"].id
    r1 = cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/analisar",
                      files=_upload()).json()
    # 1ª vez: todos os 3 alunos são novos; nenhuma turma da planilha existe ainda
    assert r1["alunos_novos"] == 3 and r1["alunos_atualizar"] == 0
    assert r1["turmas_existentes"] == 0

    assert cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar",
                        files=_upload()).status_code == 200

    r2 = cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/analisar",
                      files=_upload()).json()
    # 2ª vez (idempotente): os mesmos alunos já existem → a atualizar; turmas idem
    assert r2["alunos_novos"] == 0 and r2["alunos_atualizar"] == 3
    assert r2["turmas_existentes"] == 2


def test_confirmar_cria_turmas_alunos_matriculas_e_ficha(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    r = cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar",
                     files=_upload())
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["turmas_criadas"] == 2
    assert corpo["alunos_criados"] == 3

    turma = db.execute(select(Turma).where(Turma.escola_id == escola_id,
                                           Turma.nome == "1º Ano A")).scalar_one()
    assert turma.turno == "tarde"
    aluno = db.execute(select(Aluno).where(
        Aluno.nome == "AGATHA VITORIA MOURA DA SILVA")).scalar_one()
    assert aluno.numero_chamada == 1
    assert aluno.data_nascimento == date(2019, 9, 18)
    assert aluno.ficha == {"ra": "123.100.026-0"}   # minimização LGPD: só RA
    assert "cpf" not in aluno.ficha and "responsavel" not in aluno.ficha
    # Matriculado na turma certa, no ano ativo.
    mat = db.execute(select(Matricula).where(Matricula.aluno_id == aluno.id)).scalar_one()
    assert mat.turma_id == turma.id

    # A ficha aparece no perfil (gestor), minimizada: só o RA, sem sensíveis.
    perfil = cliente.get(f"/api/v1/escolas/{escola_id}/alunos/{aluno.id}/perfil").json()
    assert perfil["ficha"] == {"ra": "123.100.026-0"}


def _planilha_com_duplicado() -> bytes:
    """Mesmo aluno (mesmo RA) listado em DUAS turmas — a escola às vezes faz
    isso com quem está trocando de sala."""
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "1º Ano A"
    _aba(a, "1º Ano A", "TARDE", "PAULA", 300396614, [
        [1, 4446, "999.111.222-3", "OLIVER VALENTIN DE BARROS",
         date(2019, 1, 1), "MAE", "RUA A, 1", "CENTRO",
         "12-90000-0000", "", "", "", "M", "PARDA", "NÃO"],
    ])
    b = wb.create_sheet("1º Ano B")
    _aba(b, "1º ANO B", "TARDE", "MAIARA", 300396635, [
        [7, 4470, "999.111.222-3", "OLIVER VALENTIN DE BARROS",
         date(2019, 1, 1), "MAE", "RUA A, 1", "CENTRO",
         "12-90000-0000", "", "", "", "M", "PARDA", "NÃO"],
    ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_aluno_em_duas_turmas_cadastra_uma_vez_na_primeira(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    conteudo = _planilha_com_duplicado()
    files = {"arquivo": ("dup.xlsx", conteudo, CT_XLSX)}

    # A prévia já avisa o gestor e conta o aluno DISTINTO uma vez.
    ana = cliente.post(f"/api/v1/escolas/{escola_id}/importacoes/matriculas/analisar",
                       files=files).json()
    assert ana["total_alunos"] == 1          # 1 aluno distinto (mesmo RA)
    assert ana["total_registros"] == 2       # 2 linhas na planilha
    assert any("mais de uma turma" in av for av in ana["avisos"])

    base = f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar"
    r1 = cliente.post(base, files=files)
    assert r1.status_code == 200, r1.text
    assert r1.json()["alunos_criados"] == 1        # UM cadastro, não dois
    # A 2ª ocorrência é o MESMO aluno (já alocado) → não reconta como atualizado.
    assert r1.json()["alunos_atualizados"] == 0

    db.expire_all()
    oliver = db.execute(select(Aluno).where(
        Aluno.nome == "OLIVER VALENTIN DE BARROS")).scalars().all()
    assert len(oliver) == 1                          # sem duplicar
    mats = db.execute(select(Matricula).where(
        Matricula.aluno_id == oliver[0].id, Matricula.ano_letivo == 2026)).scalars().all()
    assert len(mats) == 1                            # uma matrícula só
    turma_a = db.execute(select(Turma).where(
        Turma.escola_id == escola_id, Turma.nome == "1º Ano A")).scalar_one()
    assert mats[0].turma_id == turma_a.id            # primeira turma vence

    # Reimportar mantém a MESMA turma (determinístico, não migra para a 2ª).
    assert cliente.post(base, files=files).status_code == 200
    db.expire_all()
    mats2 = db.execute(select(Matricula).where(
        Matricula.aluno_id == oliver[0].id, Matricula.ano_letivo == 2026)).scalars().all()
    assert len(mats2) == 1 and mats2[0].turma_id == turma_a.id


def test_reimportar_atualiza_sem_duplicar(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    base = f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar"
    assert cliente.post(base, files=_upload()).status_code == 200
    r2 = cliente.post(base, files=_upload())
    assert r2.status_code == 200, r2.text
    assert r2.json()["turmas_criadas"] == 0      # já existem
    assert r2.json()["alunos_criados"] == 0      # casados por RA
    assert r2.json()["alunos_atualizados"] == 3
    # Uma Agatha só (sem duplicar).
    n = db.execute(select(func.count()).select_from(Aluno).where(
        Aluno.nome == "AGATHA VITORIA MOURA DA SILVA")).scalar_one()
    assert n == 1


# --------------------------------------------------------------------------
# Atualização INCREMENTAL da Lista Piloto: quem some vira "fora_lista_piloto"
# --------------------------------------------------------------------------

def _planilha_piloto(nome_turma, alunos):
    """Planilha de UMA turma com `alunos` = [(nº, ra, nome), ...]."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nome_turma
    linhas = [[num, 4000 + i, ra, nome, date(2018, 1, 1), "RESP", "RUA X",
               "BAIRRO", "", "", "", "", "M", "PARDA", "SIM"]
              for i, (num, ra, nome) in enumerate(alunos)]
    _aba(ws, nome_turma, "TARDE", "PAULA", 300396614, linhas)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _files(conteudo, nome="Lista.xlsx"):
    return {"arquivo": (nome, conteudo, CT_XLSX)}


def _aluno(db, nome):
    return db.execute(select(Aluno).where(Aluno.nome == nome)).scalar_one()


def test_reimportar_marca_ausentes_como_fora_da_lista_sem_apagar(
        cliente, db, escola_completa):
    """Atualização incremental: aluno que some da nova lista vira
    'fora_lista_piloto' (continua no banco, nada é apagado); reimportar com ele
    de volta REATIVA; reimportar a MESMA lista não marca ninguém (idempotente)."""
    escola_id = escola_completa["escola"].id
    base = f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar"

    lista_cheia = _planilha_piloto("2º Ano A", [
        (1, "111-1", "ANA PILOTO UM"),
        (2, "222-2", "BRUNO PILOTO DOIS"),
        (3, "333-3", "CARLA PILOTO TRES")])

    # 1ª importação: cria os 3 e estabelece a associação com o piloto.
    r1 = cliente.post(base, files=_files(lista_cheia)).json()
    assert r1["alunos_criados"] == 3 and r1["alunos_fora_lista"] == 0

    # 2ª importação SEM Bruno e Carla → marcados fora, NÃO apagados.
    lista_reduzida = _planilha_piloto("2º Ano A", [(1, "111-1", "ANA PILOTO UM")])
    r2 = cliente.post(base, files=_files(lista_reduzida))
    assert r2.status_code == 200, r2.text
    assert r2.json()["alunos_fora_lista"] == 2
    assert "fora da lista piloto" in r2.json()["mensagem"]

    db.expire_all()
    assert db.execute(select(func.count()).select_from(Aluno).where(
        Aluno.nome.in_(["BRUNO PILOTO DOIS", "CARLA PILOTO TRES"]))).scalar_one() == 2
    bruno = _aluno(db, "BRUNO PILOTO DOIS")
    assert bruno.status == "fora_lista_piloto" and bruno.da_lista_piloto is True
    # Ana continua ativa (segue na lista).
    assert _aluno(db, "ANA PILOTO UM").status == "ativo"
    # Aluno da FIXTURE (não veio do piloto) permanece intacto.
    fixture = _aluno(db, "Ana Beatriz Souza")
    assert fixture.status == "ativo" and fixture.da_lista_piloto is False

    # 3ª importação com todos de volta → Bruno e Carla REATIVADOS.
    r3 = cliente.post(base, files=_files(lista_cheia))
    assert r3.status_code == 200
    assert r3.json()["alunos_fora_lista"] == 0
    db.expire_all()
    assert _aluno(db, "BRUNO PILOTO DOIS").status == "ativo"
    assert _aluno(db, "CARLA PILOTO TRES").status == "ativo"

    # 4ª importação IDÊNTICA → ninguém sai (idempotente).
    r4 = cliente.post(base, files=_files(lista_cheia))
    assert r4.json()["alunos_fora_lista"] == 0
    assert r4.json()["alunos_criados"] == 0


def test_fora_da_lista_nao_toca_aluno_manual_nem_de_upload(
        cliente, db, escola_completa):
    """A reconciliação só afeta alunos do piloto (da_lista_piloto=True). Um
    aluno criado à mão OU por upload (Matific/Elefante) matriculado no ano nunca
    é marcado 'fora_lista_piloto', mesmo ausente da lista."""
    escola_id = escola_completa["escola"].id
    base = f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar"

    # importa uma lista piloto (estabelece o piloto)
    cliente.post(base, files=_files(_planilha_piloto("2º Ano A", [
        (1, "111-1", "ANA PILOTO UM"), (2, "222-2", "BRUNO PILOTO DOIS")])))
    # cadastra um aluno MANUAL (da_lista_piloto=False por padrão) na mesma turma
    turma_id = db.execute(select(Turma.id).where(
        Turma.escola_id == escola_id, Turma.nome == "2º Ano A")).scalar_one()
    manual = cliente.post(f"/api/v1/escolas/{escola_id}/alunos",
                          json={"nome": "MANUEL CADASTRO MANUAL", "turma_id": turma_id})
    assert manual.status_code == 201, manual.text

    # nova lista SEM Bruno e SEM o manual → só Bruno (piloto) sai.
    r = cliente.post(base, files=_files(_planilha_piloto("2º Ano A", [
        (1, "111-1", "ANA PILOTO UM")]))).json()
    assert r["alunos_fora_lista"] == 1        # só o Bruno (piloto)
    db.expire_all()
    assert _aluno(db, "BRUNO PILOTO DOIS").status == "fora_lista_piloto"
    assert _aluno(db, "MANUEL CADASTRO MANUAL").status == "ativo"   # intacto


def test_reconciliacao_escopada_as_turmas_do_arquivo(cliente, db, escola_completa):
    """Subir a planilha de UMA turma NÃO tira da lista os alunos de OUTRA turma
    que nem foi enviada — a reconciliação é escopada às turmas do arquivo."""
    escola_id = escola_completa["escola"].id
    base = f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar"

    # DUAS turmas piloto, importadas separadamente.
    cliente.post(base, files=_files(_planilha_piloto("2º Ano A", [
        (1, "a1", "ANA A UM"), (2, "a2", "BRUNO A DOIS")])))
    cliente.post(base, files=_files(_planilha_piloto("2º Ano B", [
        (1, "b1", "CARLA B UM"), (2, "b2", "DAVI B DOIS")])))

    # Reimporta SÓ a turma A, sem Bruno. Só Bruno (turma A) sai; a turma B, que
    # nem foi enviada, permanece 100% intacta.
    r = cliente.post(base, files=_files(_planilha_piloto("2º Ano A", [
        (1, "a1", "ANA A UM")]))).json()
    assert r["alunos_fora_lista"] == 1
    db.expire_all()
    assert _aluno(db, "BRUNO A DOIS").status == "fora_lista_piloto"
    assert _aluno(db, "ANA A UM").status == "ativo"
    assert _aluno(db, "CARLA B UM").status == "ativo"     # turma B não enviada
    assert _aluno(db, "DAVI B DOIS").status == "ativo"


# --------------------------------------------------------------------------
# Casamento com cadastros ANTIGOS (uploads Matific/Elefante) → nome completo
# --------------------------------------------------------------------------

def _turma(db, escola_id, nome, ano_escolar):
    t = Turma(escola_id=escola_id, nome=nome, ano_escolar=ano_escolar, ano_letivo=2026)
    db.add(t); db.flush(); db.commit()
    return t


def _antigo(db, escola_id, nome, turma, *, ra="", nasc=None, com_upload=True):
    """Cadastro como os criados por upload de relatório: nome ABREVIADO, sem RA,
    matriculado numa turma, com um snapshot Matific (salvo com_upload=False)."""
    aluno = Aluno(escola_id=escola_id, nome=nome, status="ativo",
                  data_nascimento=nasc, ficha=({"ra": ra} if ra else {}))
    db.add(aluno); db.flush()
    db.add(Matricula(escola_id=escola_id, aluno_id=aluno.id, turma_id=turma.id,
                     ano_letivo=2026))
    if com_upload:
        imp = Importacao(escola_id=escola_id, plataforma="matific", tipo="pdf")
        db.add(imp); db.flush()
        db.add(SnapshotMatific(escola_id=escola_id, aluno_id=aluno.id,
                               importacao_id=imp.id, atividades=10, estrelas=5))
    db.commit()
    return aluno


def _linha(nome, ra, nasc):
    """Uma linha de aluno na ordem de COLUNAS."""
    return [1, 4000, ra, nome, nasc, "MAE", "RUA X", "BAIRRO",
            "12-90000-0000", "", "", "", "F", "PARDA", "NÃO"]


def _planilha_turma(turma_nome, linhas_alunos, meta_nome=None):
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = turma_nome
    _aba(a, meta_nome or turma_nome, "TARDE", "PROF", 300000000, linhas_alunos)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def _confirmar(cliente, escola_id, conteudo):
    return cliente.post(
        f"/api/v1/escolas/{escola_id}/importacoes/matriculas/confirmar",
        files={"arquivo": ("x.xlsx", conteudo, CT_XLSX)})


def test_predicado_abreviado_posicional():
    T = tokens_nome
    # Casos que DEVEM casar (abreviação posicional real do Matific/Elefante).
    assert casa_abreviado_posicional(T("AGATHA V"), T("AGATHA VITORIA MOURA DA SILVA"))
    assert casa_abreviado_posicional(T("AGATHA VITORIA M"), T("AGATHA VITORIA MOURA DA SILVA"))
    assert casa_abreviado_posicional(T("ELOA S"), T("ELOA SILVEIRA VASSAO"))
    assert casa_abreviado_posicional(T("PEDRO L"), T("PEDRO LUCAS SANTOS"))
    # NÃO pode casar por subsequência nem por primeiro nome diferente.
    assert not casa_abreviado_posicional(T("ELOA S"), T("ELOA VITORIA DA SILVA MARADEI"))
    assert not casa_abreviado_posicional(T("PEDRO L"), T("PEDRO ARTHUR DOURADO"))
    assert not casa_abreviado_posicional(T("ANA E"), T("MARIANA EMANUELLY"))
    # Nome idêntico vai pelo caminho nome-exato, não pelo abreviado.
    assert not casa_abreviado_posicional(
        T("AGATHA VITORIA MOURA DA SILVA"), T("AGATHA VITORIA MOURA DA SILVA"))


def test_casa_antigo_abreviado_recebe_nome_completo_e_preserva_uploads(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    antigo = _antigo(db, escola_id, "AGATHA V", turma)
    aid = antigo.id

    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "123.100.026-0", date(2019, 9, 18))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 0        # casou o antigo, não criou

    db.expire_all()
    agathas = db.execute(select(Aluno).where(
        Aluno.escola_id == escola_id, Aluno.nome.like("AGATHA%"))).scalars().all()
    assert len(agathas) == 1                       # não duplicou
    atual = db.get(Aluno, aid)
    assert atual.nome == "AGATHA VITORIA MOURA DA SILVA"   # nome completo do Excel
    assert atual.ficha["ra"] == "123.100.026-0"
    # Upload preservado — mesmo aluno_id.
    snaps = db.execute(select(func.count()).select_from(SnapshotMatific).where(
        SnapshotMatific.aluno_id == aid)).scalar_one()
    assert snaps == 1
    # Vínculo auditado.
    ev = db.execute(select(LogAuditoria).where(
        LogAuditoria.acao == "aluno.identidade_vinculada",
        LogAuditoria.entidade_id == aid)).scalars().first()
    assert ev is not None and ev.detalhes["nome_antigo"] == "AGATHA V"


def test_ambiguidade_cria_novo_e_avisa(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    a1 = _antigo(db, escola_id, "AGATHA V", turma)
    a2 = _antigo(db, escola_id, "AGATHA V", turma)
    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "111.111.111-1", date(2019, 9, 18))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 1        # ambíguo → cria novo
    assert any("Fundir alunos" in av for av in r.json()["avisos"])
    db.expire_all()
    assert db.get(Aluno, a1.id).nome == "AGATHA V"   # antigos intactos
    assert db.get(Aluno, a2.id).nome == "AGATHA V"


def test_veto_nascimento_nao_casa(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    antigo = _antigo(db, escola_id, "AGATHA V", turma, nasc=date(2018, 3, 1))
    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "111.111.111-1", date(2019, 7, 10))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 1        # nascimento diverge → não casa
    db.expire_all()
    assert db.get(Aluno, antigo.id).nome == "AGATHA V"


def test_ra_precede_abreviado(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    cheio = _antigo(db, escola_id, "AGATHA VITORIA MOURA DA SILVA", turma,
                    ra="123.100.026-0", com_upload=False)
    abrev = _antigo(db, escola_id, "AGATHA V", turma)     # sem RA, com upload
    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "123.100.026-0", date(2019, 9, 18))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 0        # casou o cheio por RA
    db.expire_all()
    assert db.get(Aluno, abrev.id).nome == "AGATHA V"     # abreviado fica intacto


def test_turma_divergente_nao_casa(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma_b = _turma(db, escola_id, "5º Ano B", "5º Ano")
    antigo = _antigo(db, escola_id, "AGATHA V", turma_b)
    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "111.111.111-1", date(2019, 9, 18))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 1        # overlap de turma < 2 → não casa
    db.expire_all()
    assert db.get(Aluno, antigo.id).nome == "AGATHA V"


def test_sem_upload_fora_do_pool_cria_novo(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    antigo = _antigo(db, escola_id, "AGATHA V", turma, com_upload=False)
    conteudo = _planilha_turma("1º Ano A", [
        _linha("AGATHA VITORIA MOURA DA SILVA", "111.111.111-1", date(2019, 9, 18))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 1        # sem upload → fora do pool
    db.expire_all()
    assert db.get(Aluno, antigo.id).nome == "AGATHA V"


def test_placeholder_ra_nao_colapsa_alunos_distintos(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    conteudo = _planilha_turma("1º Ano A", [
        _linha("JOAO PEDRO SILVA", "0", date(2019, 1, 1)),
        [2, 4001, "0", "MARIA SOUZA LIMA", date(2019, 2, 2), "MAE", "RUA", "B",
         "12-9", "", "", "", "F", "PARDA", "NÃO"],
    ])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 2        # RA "0" placeholder não colapsa
    db.expire_all()
    nomes = {a.nome for a in db.execute(select(Aluno).where(
        Aluno.escola_id == escola_id)).scalars()}
    assert "JOAO PEDRO SILVA" in nomes and "MARIA SOUZA LIMA" in nomes


def test_overlap_turma_nao_une_anos_diferentes(cliente, db, escola_completa):
    """Um antigo com matrículas em '2º Ano A' (2026) e '3º Ano B' (2025) NÃO pode
    casar com 'LUCAS SANTOS' importado em '2º Ano B' — nenhuma turma isolada tem
    série+letra em comum; só a UNIÃO das duas fabricaria o overlap (bug)."""
    escola_id = escola_completa["escola"].id
    t2a = _turma(db, escola_id, "2º Ano A", "2º Ano")
    t3b_2025 = Turma(escola_id=escola_id, nome="3º Ano B", ano_escolar="3º Ano",
                     ano_letivo=2025)
    db.add(t3b_2025); db.flush()
    lucas = Aluno(escola_id=escola_id, nome="LUCAS S", status="ativo", ficha={})
    db.add(lucas); db.flush()
    db.add(Matricula(escola_id=escola_id, aluno_id=lucas.id, turma_id=t2a.id,
                     ano_letivo=2026))
    db.add(Matricula(escola_id=escola_id, aluno_id=lucas.id, turma_id=t3b_2025.id,
                     ano_letivo=2025))
    imp = Importacao(escola_id=escola_id, plataforma="matific", tipo="pdf")
    db.add(imp); db.flush()
    db.add(SnapshotMatific(escola_id=escola_id, aluno_id=lucas.id, importacao_id=imp.id))
    db.commit()

    conteudo = _planilha_turma("2º Ano B", [
        _linha("LUCAS SANTOS OLIVEIRA", "333.333.333-3", date(2018, 5, 5))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 1        # não casou (anos/turmas distintos)
    db.expire_all()
    assert db.get(Aluno, lucas.id).nome == "LUCAS S"


def test_soft_excluido_reativa_por_ra(cliente, db, escola_completa):
    escola_id = escola_completa["escola"].id
    turma = _turma(db, escola_id, "1º Ano A", "1º Ano")
    excl = _antigo(db, escola_id, "PEDRO LUCAS SANTOS", turma,
                   ra="222.222.222-2", com_upload=False)
    excl.status = "excluido"; db.commit()
    conteudo = _planilha_turma("1º Ano A", [
        _linha("PEDRO LUCAS SANTOS", "222.222.222-2", date(2019, 3, 3))])
    r = _confirmar(cliente, escola_id, conteudo)
    assert r.status_code == 200, r.text
    assert r.json()["alunos_criados"] == 0        # casou por RA e reativou
    db.expire_all()
    assert db.get(Aluno, excl.id).status == "ativo"
    pedros = db.execute(select(func.count()).select_from(Aluno).where(
        Aluno.escola_id == escola_id, Aluno.nome.like("PEDRO%"))).scalar_one()
    assert pedros == 1                            # não duplicou
