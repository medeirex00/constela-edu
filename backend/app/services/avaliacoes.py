"""Avaliações externas: catálogo, leitura de planilha e importação de resultados.

MVP = IMPORTAÇÃO DE ARQUIVO OFICIAL (não há API oficial em nenhuma das fontes). O
importador é robusto de propósito: lê o arquivo, devolve a grade crua para a tela
deixar o gestor escolher a LINHA do cabeçalho/dados e MAPEAR as colunas por índice
(planilhas do INEP/SEDUC têm pré-âmbulo e nomes de coluna que mudam entre edições).
O casamento com a escola é pelo CÓDIGO INEP — nunca pelo nome. Só grava os níveis
que o arquivo fornece; o resto fica nulo (nunca inventa).
"""
from __future__ import annotations

import csv
import io
import ipaddress
import socket
from datetime import datetime, timedelta, timezone
from itertools import islice
from typing import Callable, Iterator
from urllib.parse import urlparse

import httpx
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import AvaliacaoExterna, Escola, FonteAvaliacao, ResultadoAvaliacao

# Teto rígido de linhas processadas na importação — barra decompression-bomb de
# XLSX (arquivo pequeno comprimido que declara milhões de linhas). Muito acima de
# qualquer arquivo real por escola (o país tem ~200 mil escolas).
_MAX_LINHAS_IMPORT = 1_000_000

# Catálogo inicial (extensível: adicionar aqui + um mapeamento na tela). O IDEB
# entra como INDICADOR (derivado), separado das avaliações de desempenho.
CATALOGO = {
    "saeb": {"nome": "SAEB", "tipo": "avaliacao", "orgao": "INEP",
             "descricao": "Sistema de Avaliação da Educação Básica (proficiência em escala SAEB)."},
    "ideb": {"nome": "IDEB", "tipo": "indicador", "orgao": "INEP",
             "descricao": "Índice derivado: desempenho (SAEB) × fluxo (aprovação). Não é medição direta."},
    "saresp": {"nome": "SARESP", "tipo": "avaliacao", "orgao": "SEDUC-SP",
               "descricao": "Avaliação de rendimento da rede estadual de São Paulo."},
    "crianca_alfabetizada": {"nome": "Criança Alfabetizada", "tipo": "avaliacao", "orgao": "INEP / CAEd",
                             "descricao": "Avaliação de alfabetização (2º ano do EF)."},
    # CAEd como FONTE/tipo próprio (decisão do dono 2026-08-04): as avaliações do CAEd
    # entram como avaliações SEPARADAS, sem alterar "Criança Alfabetizada" (que segue
    # como está). Ainda SEM preset de import de 1 clique — o mapeamento de colunas
    # será travado com o ARQUIVO OFICIAL de cada avaliação (não inventar formato); a
    # tela já permite importar por mapeamento manual quando o arquivo existir.
    "caed_fluencia": {"nome": "CAEd — Fluência Leitora", "tipo": "avaliacao", "orgao": "CAEd",
                      "descricao": "Fluência leitora aplicada pelo CAEd. Mapeamento de "
                                   "importação a configurar com o arquivo oficial."},
    "caed_diagnostica": {"nome": "CAEd — Avaliação Diagnóstica", "tipo": "avaliacao", "orgao": "CAEd",
                         "descricao": "Avaliação diagnóstica do CAEd. Mapeamento de "
                                      "importação a configurar com o arquivo oficial."},
}


def obter_avaliacao(db: Session, chave: str) -> AvaliacaoExterna:
    """Linha do catálogo para a chave (cria sob demanda a partir de CATALOGO).
    Assim funciona igual em create_all (testes), dev e produção, sem seed."""
    av = db.execute(
        select(AvaliacaoExterna).where(AvaliacaoExterna.chave == chave)
    ).scalar_one_or_none()
    if av is not None:
        return av
    meta = CATALOGO.get(chave, {"nome": chave, "tipo": "avaliacao", "orgao": None, "descricao": None})
    av = AvaliacaoExterna(chave=chave, nome=meta["nome"], tipo=meta["tipo"],
                          orgao=meta.get("orgao"), descricao=meta.get("descricao"))
    db.add(av)
    db.flush()
    return av


# --- Leitura de planilha (XLSX/CSV) -----------------------------------------

def _nao_vazia(linha: list) -> bool:
    return any(c is not None and str(c).strip() for c in linha)


# Teto do arquivo INTERNO descomprimido de um ZIP — barra zip-bomb (os arquivos
# oficiais por escola têm poucos MB; 150MB é folga larga).
_MAX_ZIP_INTERNO = 150 * 1024 * 1024


def _iterar_grade(conteudo: bytes, nome: str) -> Iterator[list]:
    """GERA as linhas não-vazias do arquivo (1ª aba do XLSX, ou CSV) em STREAMING
    — não materializa a grade inteira (senão um XLSX comprimido explodiria em
    memória). Arquivos oficiais do INEP vêm em ZIP: extrai o 1º XLSX/CSV de dentro.
    Nunca levanta por conteúdo: arquivo ilegível → não gera nada."""
    nome_l = (nome or "").lower()
    if nome_l.endswith(".zip"):
        import zipfile
        try:
            with zipfile.ZipFile(io.BytesIO(conteudo)) as zf:
                interno = next(
                    (n for n in zf.namelist()
                     if n.lower().endswith((".xlsx", ".xlsm", ".csv"))
                     and not n.startswith("__MACOSX")),
                    None)
                if interno is None or zf.getinfo(interno).file_size > _MAX_ZIP_INTERNO:
                    return
                dados = zf.read(interno)
        except Exception:  # noqa: BLE001 — zip corrompido: não gera nada
            return
        yield from _iterar_grade(dados, interno)  # recursa com o arquivo de dentro
        return
    try:
        if nome_l.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
            try:
                for row in wb.active.iter_rows(values_only=True):
                    linha = list(row)
                    if _nao_vazia(linha):
                        yield linha
            finally:
                wb.close()
        else:
            texto = conteudo.decode("utf-8-sig", errors="replace")
            delim = ";" if texto[:8192].count(";") >= texto[:8192].count(",") else ","
            for r in csv.reader(io.StringIO(texto), delimiter=delim):
                if _nao_vazia(r):
                    yield list(r)
    except Exception:  # noqa: BLE001 — arquivo corrompido/inesperado: para de gerar
        return


def analisar_planilha(conteudo: bytes, nome: str, amostra: int = 20) -> dict:
    """Grade CRUA das primeiras linhas — a tela usa para o gestor escolher a linha
    dos dados e mapear as colunas (por índice). Robusto a pré-âmbulo/edição. Lê só
    ``amostra`` linhas (não a grade inteira) — barato mesmo com arquivo enorme."""
    grade = list(islice(_iterar_grade(conteudo, nome), amostra))
    n_colunas = max((len(r) for r in grade), default=0)
    def _cel(v):
        return "" if v is None else str(v).strip()
    return {
        "linhas_lidas": len(grade),   # amostra (pode haver mais linhas no arquivo)
        "n_colunas": n_colunas,
        "primeiras_linhas": [[_cel(c) for c in r] for r in grade],
    }


# --- Normalização ------------------------------------------------------------

def normalizar_inep(valor) -> str | None:
    """Código INEP em 8 dígitos, ou None. Tolera int-como-float do openpyxl
    ('35012345.0') e separadores; NUNCA levanta."""
    if valor is None:
        return None
    s = str(valor).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = "".join(ch for ch in s if ch.isdigit())
    if not s:
        return None
    return s.zfill(8)[-8:]


def _num(valor):
    """Número a partir de célula (vírgula decimal, %, marcadores de vazio)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace("%", "").replace(",", ".")
    if s.lower() in ("", "-", "--", "nd", "*", "...", "—", "s/i", "na"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _cel(linha: list, indice) -> object:
    if indice is None or indice < 0 or indice >= len(linha):
        return None
    return linha[indice]


def _dimensao(linha: list, col: int | None, fixo: str | None) -> str | None:
    """Valor de uma dimensão (etapa/componente/turma): da coluna mapeada, ou o
    valor fixo, ou None. Célula VAZIA/ausente vira NULL — nunca a string 'None'
    (senão inventaria um nível que a fonte não deu e quebraria a idempotência)."""
    if col is not None:
        v = _cel(linha, col)
        if v is None:
            return None
        return str(v).strip() or None
    if fixo:
        return str(fixo).strip() or None
    return None


# --- Importação --------------------------------------------------------------

def importar_resultados(
    db: Session, conteudo: bytes, nome: str, *,
    avaliacao_chave: str, edicao: int, indicador: str, unidade: str,
    linha_dados: int, col_inep: int, col_valor: int,
    col_etapa: int | None = None, col_componente: int | None = None,
    col_turma: int | None = None, etapa_fixa: str | None = None,
    componente_fixo: str | None = None,
    escopo_escolas: set[int] | None = None,
) -> dict:
    """Grava os resultados do arquivo, casando escola por CÓDIGO INEP.

    ``linha_dados`` = índice (0-based, sobre as linhas NÃO-vazias) da primeira
    linha de DADOS (após cabeçalho/pré-âmbulo). Colunas por índice.
    ``escopo_escolas`` = ids elegíveis (rede) ou None (global). Idempotente:
    re-importar a mesma (avaliação, edição, indicador) ATUALIZA em vez de duplicar
    (chave: escola+etapa+componente+turma). A idempotência assume imports NÃO
    concorrentes da mesma (avaliação, edição, indicador) — é uma ação de gestão,
    sob demanda; não há corrida esperada.
    """
    avaliacao = obter_avaliacao(db, avaliacao_chave)

    # Índice CÓDIGO INEP -> (escola_id, rede_id), só das escolas do escopo.
    q = select(Escola.id, Escola.rede_id, Escola.codigo_inep).where(
        Escola.codigo_inep.isnot(None))
    if escopo_escolas is not None:
        q = q.where(Escola.id.in_(escopo_escolas))
    por_inep = {}
    for eid, rid, inep in db.execute(q):
        chave = normalizar_inep(inep)
        if chave:
            por_inep[chave] = (eid, rid)

    # Resultados JÁ existentes desta (avaliação, edição, indicador) no escopo, para
    # atualizar em vez de duplicar. Chave: (escola_id, etapa, componente, turma).
    ex_q = select(ResultadoAvaliacao).where(
        ResultadoAvaliacao.avaliacao_id == avaliacao.id,
        ResultadoAvaliacao.edicao == edicao,
        ResultadoAvaliacao.indicador == indicador)
    if escopo_escolas is not None:
        ex_q = ex_q.where(ResultadoAvaliacao.escola_id.in_(escopo_escolas))
    existentes = {
        (r.escola_id, r.etapa, r.componente, r.turma): r
        for r in db.execute(ex_q).scalars()
    }

    fonte = f"{avaliacao.nome} {edicao} — {nome}"[:200]
    inseridos = atualizados = casados = nao_casados = ignorados = processadas = 0
    inep_nao_casados: set[str] = set()
    truncado = False

    # Streaming: pula o pré-âmbulo (linha_dados) e processa até o teto anti-OOM.
    for linha in islice(_iterar_grade(conteudo, nome), linha_dados, None):
        if processadas >= _MAX_LINHAS_IMPORT:
            truncado = True
            break
        processadas += 1
        inep = normalizar_inep(_cel(linha, col_inep))
        valor = _num(_cel(linha, col_valor))
        if inep is None or valor is None:
            ignorados += 1
            continue
        alvo = por_inep.get(inep)
        if alvo is None:
            nao_casados += 1
            if len(inep_nao_casados) < 50:
                inep_nao_casados.add(inep)
            continue
        escola_id, rede_id = alvo
        etapa = _dimensao(linha, col_etapa, etapa_fixa)
        componente = _dimensao(linha, col_componente, componente_fixo)
        turma = _dimensao(linha, col_turma, None)
        casados += 1

        chave = (escola_id, etapa, componente, turma)
        atual = existentes.get(chave)
        if atual is not None:
            atual.valor, atual.unidade, atual.fonte = valor, unidade, fonte
            atual.rede_id, atual.codigo_inep = rede_id, inep
            atualizados += 1
        else:
            novo = ResultadoAvaliacao(
                avaliacao_id=avaliacao.id, edicao=edicao, rede_id=rede_id,
                escola_id=escola_id, codigo_inep=inep, etapa=etapa,
                componente=componente, turma=turma, indicador=indicador,
                valor=valor, unidade=unidade, fonte=fonte)
            db.add(novo)
            existentes[chave] = novo
            inseridos += 1

    return {
        "avaliacao": avaliacao.chave,
        "edicao": edicao,
        "indicador": indicador,
        "linhas_dados": processadas,
        "truncado": truncado,
        "casados": casados,
        "inseridos": inseridos,
        "atualizados": atualizados,
        "nao_casados": nao_casados,
        "ignorados": ignorados,
        "inep_nao_casados_amostra": sorted(inep_nao_casados)[:20],
    }


# --- Correlação / evolução (cruzar avaliação × engajamento) -------------------

def _escolas_da_rede_ids(db: Session, rede_id: int) -> list[int]:
    return list(db.execute(select(Escola.id).where(Escola.rede_id == rede_id)).scalars())


def opcoes_rede(db: Session, rede_id: int) -> list[dict]:
    """As séries de avaliação que EXISTEM para as escolas da rede — por
    (avaliação, indicador, etapa, componente) e as edições presentes. A tela usa
    para montar os seletores sem oferecer combinação vazia."""
    ids = _escolas_da_rede_ids(db, rede_id)
    if not ids:
        return []
    linhas = db.execute(
        select(AvaliacaoExterna.chave, AvaliacaoExterna.nome, AvaliacaoExterna.tipo,
               ResultadoAvaliacao.indicador, ResultadoAvaliacao.etapa,
               ResultadoAvaliacao.componente, ResultadoAvaliacao.edicao)
        .join(ResultadoAvaliacao, ResultadoAvaliacao.avaliacao_id == AvaliacaoExterna.id)
        .where(ResultadoAvaliacao.escola_id.in_(ids))
        .distinct()
    ).all()
    series: dict[tuple, dict] = {}
    for chave, nome, tipo, indicador, etapa, componente, edicao in linhas:
        k = (chave, indicador, etapa, componente)
        s = series.setdefault(k, {
            "avaliacao": chave, "nome": nome, "tipo": tipo, "indicador": indicador,
            "etapa": etapa, "componente": componente, "edicoes": set()})
        s["edicoes"].add(edicao)
    saida = []
    for s in series.values():
        s = {**s, "edicoes": sorted(s["edicoes"], reverse=True)}
        saida.append(s)
    saida.sort(key=lambda s: (s["nome"], s["indicador"] or "", s["etapa"] or "",
                              s["componente"] or ""))
    return saida


def _pearson(pares: list[tuple[float, float]]) -> float | None:
    """Correlação de Pearson de uma lista de (x, y). None se n<3 ou variância nula
    (sem correlação definível — a tela mostra 'dados insuficientes')."""
    n = len(pares)
    if n < 3:
        return None
    mx = sum(x for x, _ in pares) / n
    my = sum(y for _, y in pares) / n
    sxy = sum((x - mx) * (y - my) for x, y in pares)
    sxx = sum((x - mx) ** 2 for x, _ in pares)
    syy = sum((y - my) ** 2 for _, y in pares)
    if sxx <= 0 or syy <= 0:
        return None
    return round(sxy / (sxx ** 0.5 * syy ** 0.5), 3)


def correlacao_rede(db: Session, rede_id: int, *, avaliacao_chave: str, indicador: str,
                    edicao: int, etapa: str | None = None, componente: str | None = None,
                    metrica: str = "pontuacao_geral") -> dict:
    """Cruza, por ESCOLA da rede, o valor da avaliação (edição escolhida) com o
    ENGAJAMENTO nas plataformas. Devolve os pontos do gráfico de dispersão, a
    correlação de Pearson e a EVOLUÇÃO histórica (média da rede por edição).
    NUNCA afirma causalidade — a tela deixa isso explícito.

    O eixo X padrão é ``pontuacao_geral`` (índice per capita 0–1000, comparável
    entre escolas). Com ``media_geral`` — normalizada contra o P90 de cada escola
    — a dispersão cruzava o SAEB com um número que mede a homogeneidade interna
    da escola, e o Pearson resultante não significava o que a tela dizia."""
    from app.services import rede as svc_rede  # engajamento por escola (reuso)

    if metrica not in ("pontuacao_geral", "media_geral", "adocao"):
        metrica = "pontuacao_geral"
    av = obter_avaliacao(db, avaliacao_chave)
    engaj = {c["escola_id"]: c for c in svc_rede._kpis_da_rede(db, rede_id)}
    ids = _escolas_da_rede_ids(db, rede_id)

    def _filtro(q):
        # ESCOPO da rede (isolamento): só resultados das escolas DESTA rede — senão
        # a evolução somaria escolas de outras redes que importaram a mesma série.
        q = q.where(ResultadoAvaliacao.avaliacao_id == av.id,
                    ResultadoAvaliacao.indicador == indicador,
                    ResultadoAvaliacao.escola_id.in_(ids))
        q = q.where(ResultadoAvaliacao.etapa == etapa) if etapa else q.where(
            ResultadoAvaliacao.etapa.is_(None))
        q = q.where(ResultadoAvaliacao.componente == componente) if componente else q.where(
            ResultadoAvaliacao.componente.is_(None))
        return q

    # Valor por ESCOLA (média sobre eventuais turmas) na edição escolhida — 1 ponto
    # por escola, determinístico mesmo se a fonte trouxer resultado por turma.
    valores = {
        eid: v for eid, v in db.execute(
            _filtro(select(ResultadoAvaliacao.escola_id,
                           func.avg(ResultadoAvaliacao.valor)))
            .where(ResultadoAvaliacao.edicao == edicao)
            .group_by(ResultadoAvaliacao.escola_id))
    }
    pontos = []
    for eid, valor in valores.items():
        c = engaj.get(eid)
        if c is None or c.get("alunos_com_dados", 0) <= 0:
            continue  # sem engajamento com dados: não entra na correlação
        x = float(c[metrica])
        pontos.append({"escola_id": eid, "nome": c["nome"], "x": round(x, 1),
                       "y": round(float(valor), 2)})
    pearson = _pearson([(p["x"], p["y"]) for p in pontos])

    # Evolução: média por (escola, edição) e DEPOIS média entre as escolas por
    # edição (ponderação por escola, não por linha) + contagem de escolas DISTINTAS.
    sub = (_filtro(select(ResultadoAvaliacao.edicao.label("ed"),
                          ResultadoAvaliacao.escola_id.label("eid"),
                          func.avg(ResultadoAvaliacao.valor).label("v")))
           .group_by(ResultadoAvaliacao.edicao, ResultadoAvaliacao.escola_id)
           .subquery())
    ev = db.execute(
        select(sub.c.ed, func.avg(sub.c.v), func.count(sub.c.eid))
        .group_by(sub.c.ed).order_by(sub.c.ed)
    ).all()
    evolucao = [{"edicao": e, "media_valor": round(float(m), 2), "escolas": int(n)}
                for e, m, n in ev]

    return {
        "avaliacao": av.chave, "nome": av.nome, "tipo": av.tipo,
        "indicador": indicador, "etapa": etapa, "componente": componente,
        "edicao": edicao, "metrica": metrica,
        "pontos": pontos, "n": len(pontos), "pearson": pearson,
        "evolucao": evolucao,
    }


# --- Coleta AUTOMÁTICA (o "robô": baixa o arquivo oficial e importa) ----------

Baixador = Callable[[str], bytes]
_UA_NAVEGADOR = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                 "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")
# Teto do arquivo baixado. O IDEB oficial por escola é GRANDE: o ZIP de "anos
# iniciais" 2023 tem ~97 MB (XLSX de ~50 MB dentro, 126 colunas, ~65 mil escolas).
# 150 MB dá folga sobre o real e ainda barra abuso.
_MAX_DOWNLOAD = 150 * 1024 * 1024


def _agora() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


_MAX_REDIRECTS = 5


def _exigir_destino_publico(url: str) -> None:
    """Anti-SSRF: só https e host que resolve para IP PÚBLICO (rejeita privado/
    loopback/link-local/reservado/multicast — inclui 169.254.169.254 de metadata).
    Aplicado ANTES de cada conexão, inclusive a cada redirect (revalida o salto)."""
    partes = urlparse(url)
    if partes.scheme != "https":
        raise ValueError(f"esquema não permitido (só https): {partes.scheme or 'vazio'}")
    host = partes.hostname
    if not host:
        raise ValueError("URL sem host")
    try:
        infos = socket.getaddrinfo(host, partes.port or 443, proto=socket.IPPROTO_TCP)
    except socket.gaierror as exc:
        raise ValueError(f"host não resolvido: {host}") from exc
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if (ip.is_private or ip.is_loopback or ip.is_link_local
                or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
            raise ValueError(f"host aponta para rede não-pública ({ip}): {host}")


def baixar_url(url: str, *, timeout: float = 30.0) -> bytes:
    """Baixa o arquivo oficial (User-Agent de navegador — os portais oficiais
    bloqueiam clientes 'crus'). Guarda anti-SSRF: só https para host público, e os
    redirects são seguidos MANUALMENTE revalidando cada destino (o alvo final não
    fica a cargo do upstream). LEVANTA em erro; o chamador registra o status."""
    cabecalhos = {"User-Agent": _UA_NAVEGADOR, "Accept": "*/*"}
    with httpx.Client(timeout=timeout, follow_redirects=False, headers=cabecalhos) as cli:
        for _ in range(_MAX_REDIRECTS + 1):
            _exigir_destino_publico(url)
            with cli.stream("GET", url) as r:
                if r.is_redirect:
                    destino = r.headers.get("location")
                    if not destino:
                        raise ValueError("redirect sem destino")
                    url = str(r.url.join(destino))  # resolve relativo → absoluto
                    continue
                r.raise_for_status()
                buf = bytearray()
                for chunk in r.iter_bytes():
                    buf += chunk
                    if len(buf) > _MAX_DOWNLOAD:
                        raise ValueError("arquivo excede o teto de download")
                return bytes(buf)
    raise ValueError("excesso de redirects")


def _nome_do_url(url: str) -> str:
    return url.rstrip("/").split("/")[-1] or "arquivo"


def coletar_fonte(db: Session, fonte: FonteAvaliacao,
                  baixador: Baixador | None = None) -> dict:
    """Baixa o arquivo da fonte e importa (MESMA máquina do upload manual), casando
    por INEP globalmente. NUNCA levanta: erro de download/parse vira ultimo_status=
    'erro' (o painel mostra — não quebra em silêncio). ``baixador`` é injetável
    (default = ``baixar_url``, resolvido em tempo de chamada p/ ser testável)."""
    baixador = baixador or baixar_url
    # Serializa com outra coleta simultânea da MESMA fonte (scheduler × "coletar
    # agora"): a 2ª relê os resultados já gravados e ATUALIZA em vez de duplicar.
    from app.core.database import bloquear_fonte_para_coleta
    bloquear_fonte_para_coleta(db, fonte.id)
    av = db.get(AvaliacaoExterna, fonte.avaliacao_id)
    m = fonte.mapeamento or {}
    try:
        conteudo = baixador(fonte.url)
        resumo = importar_resultados(
            db, conteudo, _nome_do_url(fonte.url),
            avaliacao_chave=av.chave, edicao=fonte.edicao,
            indicador=fonte.indicador, unidade=fonte.unidade,
            linha_dados=int(m.get("linha_dados", 0)),
            col_inep=int(m.get("col_inep", 0)), col_valor=int(m.get("col_valor", 0)),
            col_etapa=m.get("col_etapa"), col_componente=m.get("col_componente"),
            col_turma=m.get("col_turma"),
            etapa_fixa=m.get("etapa_fixa"), componente_fixo=m.get("componente_fixo"),
            escopo_escolas=None)  # global: casa qualquer escola da base por código INEP
        fonte.ultimo_status, fonte.ultimo_erro, fonte.ultimo_resumo = "ok", None, resumo
    except Exception as exc:  # noqa: BLE001 — nunca quebra: registra o erro
        fonte.ultimo_status = "erro"
        fonte.ultimo_erro = f"{type(exc).__name__}: {exc}"[:300]
        fonte.ultimo_resumo = {"erro": fonte.ultimo_erro}
    fonte.ultima_coleta = _agora()
    fonte.proxima_coleta = (_agora() + timedelta(days=30)) if fonte.cadencia == "mensal" else None
    return {"status": fonte.ultimo_status, **(fonte.ultimo_resumo or {})}


def coletar_pendentes(db: Session, baixador: Baixador | None = None,
                      limite: int | None = None) -> int:
    """Coleta as fontes ATIVAS com proxima_coleta vencida — chamado pelo scheduler.
    Devolve quantas foram coletadas. ``limite`` é o TETO por rodada (a coleta roda
    embutida na thread do sync; sem teto, muitas fontes — ou uma lenta — poderiam
    estagnar a fila). O que sobrar vai na próxima rodada; a mais antiga vai primeiro."""
    limite = settings.AVALIACOES_COLETA_POR_RODADA if limite is None else limite
    pendentes = db.execute(
        select(FonteAvaliacao).where(
            FonteAvaliacao.ativo.is_(True),
            FonteAvaliacao.proxima_coleta.isnot(None),
            FonteAvaliacao.proxima_coleta <= _agora(),
        ).order_by(FonteAvaliacao.proxima_coleta).limit(limite)
    ).scalars().all()
    for fonte in pendentes:
        coletar_fonte(db, fonte, baixador)
    if pendentes:
        db.commit()
    return len(pendentes)
