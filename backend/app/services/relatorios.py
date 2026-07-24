"""Exportação de relatórios (PRD §86–§103): CSV, Excel e PDF + certificados.

Todos os formatos partem das mesmas linhas (cabeçalho + dados), garantindo
que CSV, Excel e PDF nunca divirjam. A identidade visual (nome da escola e
cor primária configurável) aparece no Excel e no PDF.
"""
from __future__ import annotations

import base64
import csv
import io
import mimetypes
import re
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (Aluno, Escola, Leitura, Livro, Matricula, Nota,
                        SnapshotElefante, SnapshotMatific, Turma)
from app.services import scoring

COR_PRIMARIA_PADRAO = "#1B2A4A"  # azul profundo da marca Constela Edu


def cor_primaria(db: Session, escola_id: int) -> str:
    aparencia = scoring.obter_config(db, escola_id, "aparencia", "valores", {})
    cor = str(aparencia.get("cor_primaria") or COR_PRIMARIA_PADRAO)
    if not cor.startswith("#") or len(cor) != 7:
        return COR_PRIMARIA_PADRAO
    return cor


# ---------------------------------------------------------------------------
# Fontes de dados (mesmas linhas para todos os formatos)
# ---------------------------------------------------------------------------

def linhas_ranking(db: Session, escola_id: int) -> tuple[list[str], list[list]]:
    escola = db.get(Escola, escola_id)
    linhas = db.execute(
        select(Nota, Aluno, Turma)
        .join(Aluno, Nota.aluno_id == Aluno.id)
        .join(Matricula, (Matricula.aluno_id == Aluno.id)
              & (Matricula.ano_letivo == escola.ano_letivo_ativo))
        .join(Turma, Matricula.turma_id == Turma.id)
        # Aluno.status: relatório exportado não inclui alunos arquivados/excluídos.
        .where(Nota.escola_id == escola_id, Nota.ano_letivo == escola.ano_letivo_ativo,
               Aluno.status == "ativo")
        .order_by(Nota.posicao)
    ).all()
    cabecalho = ["Posição", "Aluno", "Turma", "Série", "Nota Matific",
                 "Nota Elefante", "Nota Geral"]
    return cabecalho, [
        [nota.posicao, aluno.nome, turma.nome, turma.ano_escolar,
         nota.nota_matific, nota.nota_elefante, nota.nota_geral]
        for nota, aluno, turma in linhas
    ]


def linhas_alunos(db: Session, escola_id: int) -> tuple[list[str], list[list]]:
    escola = db.get(Escola, escola_id)
    linhas = db.execute(
        select(Aluno, Turma)
        .join(Matricula, (Matricula.aluno_id == Aluno.id)
              & (Matricula.ano_letivo == escola.ano_letivo_ativo))
        .join(Turma, Matricula.turma_id == Turma.id)
        .where(Aluno.escola_id == escola_id, Aluno.status == "ativo")
        .order_by(Turma.nome, Aluno.nome)
    ).all()
    cabecalho = ["Aluno", "Turma", "Série", "Nº de chamada", "Situação"]
    return cabecalho, [
        [aluno.nome, turma.nome, turma.ano_escolar,
         aluno.numero_chamada or "", aluno.status]
        for aluno, turma in linhas
    ]


def linhas_livros(db: Session, escola_id: int) -> tuple[list[str], list[list]]:
    from sqlalchemy import func

    contagem = dict(db.execute(
        select(Leitura.livro_id, func.count(Leitura.id))
        .where(Leitura.escola_id == escola_id)
        .group_by(Leitura.livro_id)
    ).all())
    livros = db.execute(
        select(Livro).where(Livro.escola_id == escola_id).order_by(Livro.titulo)
    ).scalars().all()
    cabecalho = ["Título", "Autor", "Nível", "Categoria", "Leituras"]
    return cabecalho, [
        [livro.titulo, livro.autor or "", livro.nivel_codigo,
         livro.categoria or "", contagem.get(livro.id, 0)]
        for livro in livros
    ]


FONTES = {
    "ranking": ("Ranking Geral", linhas_ranking),
    "alunos": ("Lista de Alunos", linhas_alunos),
    "livros": ("Catálogo de Livros", linhas_livros),
}


# ---------------------------------------------------------------------------
# Geradores por formato
# ---------------------------------------------------------------------------

def gerar_csv(cabecalho: list[str], linhas: list[list]) -> bytes:
    saida = io.StringIO()
    escritor = csv.writer(saida, delimiter=";", lineterminator="\r\n")
    escritor.writerow(cabecalho)
    escritor.writerows(linhas)
    # BOM para o Excel pt-BR abrir com acentuação correta
    return b"\xef\xbb\xbf" + saida.getvalue().encode("utf-8")


def gerar_xlsx(titulo: str, escola_nome: str, cor: str,
               cabecalho: list[str], linhas: list[list]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    cor_hex = cor.lstrip("#").upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cabecalho))
    celula = ws.cell(row=1, column=1, value=f"{escola_nome} — {titulo}")
    celula.font = Font(bold=True, color="FFFFFF", size=13)
    celula.fill = PatternFill("solid", fgColor=cor_hex)
    celula.alignment = Alignment(horizontal="center")

    for coluna, nome in enumerate(cabecalho, start=1):
        celula = ws.cell(row=2, column=coluna, value=nome)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor="EEEEEE")
    for indice, linha in enumerate(linhas, start=3):
        for coluna, valor in enumerate(linha, start=1):
            ws.cell(row=indice, column=coluna, value=valor)
    for coluna in range(1, len(cabecalho) + 1):
        largura = max(
            [len(str(cabecalho[coluna - 1]))]
            + [len(str(linha[coluna - 1])) for linha in linhas[:200]]
        )
        ws.column_dimensions[ws.cell(row=2, column=coluna).column_letter].width = min(40, largura + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _hex_para_rgb(cor: str) -> tuple[int, int, int]:
    cor = cor.lstrip("#")
    return int(cor[0:2], 16), int(cor[2:4], 16), int(cor[4:6], 16)


def _latin1(texto) -> str:
    """As fontes nativas do PDF só suportam latin-1: os acentos do português
    passam intactos; qualquer outro caractere vira '?' em vez de derrubar o
    relatório inteiro com erro 500."""
    return str(texto).encode("latin-1", "replace").decode("latin-1")


def _couber(pdf, texto, largura: float) -> str:
    """Corta o texto para caber na LARGURA REAL da célula (não em nº fixo de
    letras), terminando com '...' (o '…' não existe em latin-1)."""
    texto = _latin1(texto)
    if pdf.get_string_width(texto) <= largura - 2:
        return texto
    while texto and pdf.get_string_width(texto + "...") > largura - 2:
        texto = texto[:-1]
    return texto + "..."


def _larguras_colunas(pdf, cabecalho: list[str], linhas: list[list],
                      total: float) -> list[float]:
    """Largura proporcional ao conteúdo de cada coluna (com piso), para a
    coluna de nomes — hoje com nomes COMPLETOS da lista de matrículas — não
    ficar espremida como as colunas numéricas."""
    pesos = []
    for i, nome in enumerate(cabecalho):
        maior = pdf.get_string_width(str(nome)) + 4
        for linha in linhas[:300]:
            if i < len(linha):
                maior = max(maior, pdf.get_string_width(_latin1(linha[i])) + 4)
        pesos.append(min(maior, total * 0.45))  # nenhuma coluna domina a página
    fator = total / sum(pesos)
    return [p * fator for p in pesos]


def gerar_pdf(titulo: str, escola_nome: str, cor: str,
              cabecalho: list[str], linhas: list[list]) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    r, g, b = _hex_para_rgb(cor)
    pdf.set_fill_color(r, g, b)
    pdf.rect(0, 0, 210, 22, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 14)
    pdf.set_xy(10, 6)
    pdf.cell(0, 6, _latin1(escola_nome))
    pdf.set_font("helvetica", "", 10)
    pdf.set_xy(10, 13)
    agora = datetime.now(timezone.utc).strftime("%d/%m/%Y")
    pdf.cell(0, 5, _latin1(f"{titulo} - gerado em {agora}"))

    pdf.set_y(28)
    pdf.set_text_color(30, 30, 30)
    pdf.set_font("helvetica", "", 9)
    larguras = _larguras_colunas(pdf, cabecalho, linhas, total=190)

    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(238, 238, 238)
    for nome, largura in zip(cabecalho, larguras):
        pdf.cell(largura, 7, _couber(pdf, nome, largura), border=1, fill=True)
    pdf.ln()

    pdf.set_font("helvetica", "", 9)
    for linha in linhas:
        for valor, largura in zip(linha, larguras):
            pdf.cell(largura, 6.5, _couber(pdf, valor, largura), border=1)
        pdf.ln()

    return bytes(pdf.output())


def gerar_certificado(escola_nome: str, cor: str, aluno_nome: str,
                      turma: str, posicao: int | None, nota_geral: float,
                      ano_letivo: int, logos: tuple[str, str] | None = None) -> bytes:
    """Certificado individual em paisagem (PRD §99).

    Renderiza um certificado ELEGANTE (moldura azul-marinho + dourado, logo
    Constela, medalha) via HTML→PDF (Chromium). Se o navegador não estiver
    disponível no ambiente (ex.: CI sem browser), cai para a versão fpdf simples
    — o certificado NUNCA falha por falta do Chromium."""
    try:
        return _certificado_html_pdf(escola_nome, cor, aluno_nome, turma,
                                     posicao, nota_geral, ano_letivo, logos)
    except Exception:  # noqa: BLE001 — sem Chromium/erro de render: usa a reserva
        return _certificado_fpdf(escola_nome, cor, aluno_nome, turma,
                                 posicao, nota_geral, ano_letivo)


# Logo oficial (C) embutida como SVG — a mesma de identidade/logo-oficial.png.
_LOGO_C_SVG = (
    '<svg viewBox="0 0 100 100" width="66" height="66">'
    '<rect width="100" height="100" rx="22" fill="#1B2A4A"/>'
    '<path d="M 67,20.555 A 34,34 0 1 0 67,79.445" fill="none" stroke="#FFFFFF"'
    ' stroke-width="7" stroke-linecap="round"/>'
    '<circle cx="33" cy="79.445" r="4.2" fill="#FFFFFF"/>'
    '<circle cx="16" cy="50" r="4.6" fill="#FFFFFF"/>'
    '<circle cx="33" cy="20.555" r="4.2" fill="#FFFFFF"/>'
    '<circle cx="50" cy="50" r="12.5" fill="#F5B942"/></svg>'
)
_CANTO_SVG = (
    '<svg viewBox="0 0 300 300" class="canto">'
    '<path d="M0,0 L300,0 Q140,8 78,86 Q14,168 0,300 Z" fill="#1B2A4A"/>'
    '<path d="M0,0 L232,0 Q108,12 60,74 Q14,140 0,232 Z" fill="none"'
    ' stroke="#F5B942" stroke-width="5" opacity="0.9"/></svg>'
)
_MEDALHA_SVG = (
    '<svg viewBox="0 0 72 100" width="66" height="92">'
    '<path d="M27 46 L18 82 L31 73 L36 86 L41 73 L54 82 L45 46 Z" fill="#1B2A4A"/>'
    '<circle cx="36" cy="30" r="26" fill="#F5B942" stroke="#1B2A4A" stroke-width="3"/>'
    '<circle cx="36" cy="30" r="19" fill="none" stroke="#1B2A4A" stroke-width="1.5" opacity="0.45"/>'
    '<path d="M36 15 l4.2 9.4 10.3 1 -7.7 6.9 2.4 10.1 -9.2 -5.4 -9.2 5.4 2.4 -10.1'
    ' -7.7 -6.9 10.3 -1 Z" fill="#1B2A4A"/></svg>'
)

# ---------------------------------------------------------------------------
# Logos institucionais (cidade / prefeitura) — OPCIONAIS, por arquivo em
# app/marca/. Quando existem, entram no cabeçalho dos documentos "de vitrine"
# (certificado e cartaz do ranking): brasão da cidade à esquerda, marca
# Constela ao centro, prefeitura à direita — como no exemplo do piloto de
# Caraguatatuba. Sem os arquivos, aparece só a marca Constela (o produto segue
# genérico para qualquer escola — é só soltar os PNGs da cidade na pasta).
# ---------------------------------------------------------------------------
_MARCA_DIR = Path(__file__).resolve().parent.parent / "marca"


def _asset_data_uri(nome: str) -> str:
    """data: URI de um arquivo de app/marca/ (ou '' se não existir).

    Sem cache de propósito: os arquivos são lidos a cada documento (leitura de
    poucos KB, rara — o gargalo é o Chromium), para que trocar/adicionar um logo
    na pasta tenha efeito imediato, sem precisar reiniciar o processo."""
    try:
        dados = (_MARCA_DIR / nome).read_bytes()
    except OSError:
        return ""
    mime = mimetypes.guess_type(nome)[0] or "image/png"
    return f"data:{mime};base64," + base64.b64encode(dados).decode("ascii")


def logos_da_escola(db: Session, escola_id: int) -> tuple[str, str]:
    """(brasão, prefeitura) DESTA escola como data URIs, se foram enviados em
    Configurações → Aparência. Strings vazias quando não há — aí os documentos
    caem para o logo padrão de `app/marca/` (piloto) e, na falta, só a marca
    Constela. É isto que torna o brasão/prefeitura automáticos por cidade."""
    cfg = scoring.obter_config(db, escola_id, "aparencia", "logos", {})
    return (str(cfg.get("brasao_data_uri") or ""),
            str(cfg.get("prefeitura_data_uri") or ""))


def _cabecalho_logos(logos: tuple[str, str] | None = None) -> str:
    """Cabeçalho com até 3 logos (brasão · Constela · prefeitura).

    Regra ALL-OR-NOTHING por escola: se a escola enviou QUALQUER logo próprio
    (Aparência), usa SÓ os dela — nunca completa o que falta com o padrão de
    OUTRA cidade (evita misturar, ex.: brasão da cidade nova + prefeitura de
    Caraguá). Só quando a escola não tem NENHUM logo próprio é que caímos para
    o padrão do piloto em `app/marca/`; e na falta dele, só a marca Constela."""
    brasao_uri, pref_uri = logos or ("", "")
    if brasao_uri or pref_uri:
        brasao, prefeitura = brasao_uri, pref_uri
    else:
        brasao = _asset_data_uri("cidade-brasao.png")
        prefeitura = _asset_data_uri("prefeitura.png")
    esq = f'<img class="logo-cidade" src="{brasao}" alt="Brasão da cidade">' if brasao else ""
    dire = f'<img class="logo-pref" src="{prefeitura}" alt="Prefeitura">' if prefeitura else ""
    return (f'<div class="cabecalho">{esq}'
            f'<div class="constela">{_LOGO_C_SVG}'
            f'<span class="nome">Constela <b>Edu</b></span></div>'
            f'{dire}</div>')


# Ícones (SVG dourado) das colunas de nota do cartaz do ranking.
_IC_STAR = ('<svg class="ic" width="13" height="13" viewBox="0 0 24 24">'
            '<path d="M12 2l2.9 6.26L22 9.27l-5 4.87 1.18 6.88L12 17.77 5.82 21'
            'l1.18-6.88-5-4.87 6.9-1.01z" fill="#F5B841"/></svg>')
_IC_TROFEU = ('<svg class="ic" width="13" height="13" viewBox="0 0 24 24" fill="#F5B841">'
              '<path d="M18 3H6v2H3v3a4 4 0 0 0 4 4 5 5 0 0 0 4 3v2H8v2h8v-2h-3v-2'
              'a5 5 0 0 0 4-3 4 4 0 0 0 4-4V5h-3zM5 8V7h1v3a2 2 0 0 1-1-2zm14 0'
              'a2 2 0 0 1-1 2V7h1z"/></svg>')
_IC_ELEFANTE = ('<svg class="ic" width="15" height="14" viewBox="0 0 34 32" fill="#F5B841">'
                '<ellipse cx="8" cy="14" rx="7" ry="9"/>'
                '<ellipse cx="26" cy="14" rx="7" ry="9"/>'
                '<circle cx="17" cy="14" r="10.5"/>'
                '<path d="M13.5 17h7v6.5a3 3 0 0 0 6 0v-1.2a1.1 1.1 0 0 0-2.2 0v1.2'
                'a0.8 0.8 0 0 1-1.6 0V17z"/></svg>')


def _fmt_nota(valor) -> str:
    """Nota 0–100 com uma casa e vírgula decimal ('85,3'); '—' se vazia."""
    try:
        return f"{float(valor):.1f}".replace(".", ",")
    except (TypeError, ValueError):
        return "—"


_CERT_TEMPLATE = """<style>
  @page { size: A4 landscape; margin: 0; }
  * { margin:0; padding:0; box-sizing:border-box; }
  html,body { width:297mm; height:210mm; -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  body { font-family: Georgia,'Times New Roman',serif; background:#ffffff; position:relative; overflow:hidden; }
  .canto { position:absolute; width:118mm; height:118mm; }
  .c-tl { top:0; left:0; }
  .c-tr { top:0; right:0; transform:scaleX(-1); }
  .c-bl { bottom:0; left:0; transform:scaleY(-1); }
  .c-br { bottom:0; right:0; transform:scale(-1,-1); }
  .frame { position:absolute; inset:11mm; border:2px solid #1B2A4A; }
  .frame::after { content:""; position:absolute; inset:3.5mm; border:1px solid #C9A24B; }
  .conteudo { position:absolute; inset:0; display:flex; flex-direction:column; align-items:center;
              justify-content:flex-start; text-align:center; padding:20mm 40mm 0; }
  .cabecalho { display:flex; align-items:center; justify-content:center; gap:44px; }
  .cabecalho .logo-cidade { height:74px; width:74px; }
  .cabecalho .logo-pref { height:60px; width:auto; }
  .cabecalho .constela { display:flex; flex-direction:column; align-items:center; gap:4px; }
  .cabecalho .constela .nome { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-weight:800;
                 font-size:19px; letter-spacing:1px; color:#1B2A4A; }
  .cabecalho .constela .nome b { color:#F5B942; }
  h1 { font-size:52px; font-weight:700; letter-spacing:8px; color:#1B2A4A; margin-top:12px; }
  .divisor { display:flex; align-items:center; gap:12px; margin:8px 0 14px; color:#C9A24B; }
  .divisor .linha { width:120px; height:1.5px; background:#C9A24B; }
  .divisor .losango { width:9px; height:9px; background:#F5B942; transform:rotate(45deg); }
  .intro { font-size:15px; color:#555; letter-spacing:.5px; }
  .aluno { font-size:40px; font-weight:700; color:#1B2A4A; margin:14px 0 6px; letter-spacing:.5px; text-transform:uppercase; }
  .sublinhado { width:60%; max-width:520px; height:1.5px; background:#C9A24B; margin:0 auto 16px; }
  .detalhe { font-size:15.5px; color:#444; line-height:1.7; max-width:640px; }
  .detalhe b { color:#1B2A4A; }
  .medalha { margin:14px 0 4px; }
  .emitido { font-size:12.5px; color:#666; letter-spacing:.5px; }
  .assinatura { position:absolute; bottom:26mm; left:50%; transform:translateX(-50%); text-align:center; }
  .assinatura .linha { width:230px; height:1px; background:#8a8a8a; margin:0 auto 6px; }
  .assinatura .rotulo { font-size:12.5px; color:#555; }
</style>
⟦CANTOS⟧
<div class="frame"></div>
<div class="conteudo">
  ⟦CABECALHO⟧
  <h1>CERTIFICADO</h1>
  <div class="divisor"><span class="linha"></span><span class="losango"></span><span class="linha"></span></div>
  <p class="intro">A escola ⟦ESCOLA⟧ certifica que</p>
  <p class="aluno">⟦ALUNO⟧</p>
  <div class="sublinhado"></div>
  <p class="detalhe">da turma <b>⟦TURMA⟧</b>, alcançou a nota geral <b>⟦NOTA⟧</b>⟦POSICAO⟧<br>
     no <b>Constela Edu</b> — ano letivo de <b>⟦ANO⟧</b>.</p>
  <div class="medalha">⟦MEDALHA⟧</div>
  <p class="emitido">Emitido em <b>⟦DATA⟧</b></p>
</div>
<div class="assinatura"><div class="linha"></div><div class="rotulo">Direção</div></div>
"""


def _esc_html(texto: str) -> str:
    # Escapa HTML e NEUTRALIZA os delimitadores de placeholder ⟦ ⟧: sem isso,
    # um nome/escola que contivesse literalmente "⟦LINHAS⟧" reintroduziria um
    # token que um .replace posterior preencheria, corrompendo o documento.
    return (str(texto).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace("⟦", "").replace("⟧", ""))


def _certificado_html_pdf(escola_nome: str, cor: str, aluno_nome: str,
                          turma: str, posicao: int | None, nota_geral: float,
                          ano_letivo: int, logos: tuple[str, str] | None = None) -> bytes:
    """Renderiza o certificado bonito em PDF via Chromium (Playwright)."""
    from playwright.sync_api import sync_playwright

    cantos = "".join(
        _CANTO_SVG.replace('class="canto"', f'class="canto {c}"')
        for c in ("c-tl", "c-tr", "c-bl", "c-br"))
    pos_txt = (f" e a <b>{posicao}ª posição</b> no Ranking Geral" if posicao else "")
    html = (_CERT_TEMPLATE
            .replace("⟦CANTOS⟧", cantos)
            .replace("⟦CABECALHO⟧", _cabecalho_logos(logos))
            .replace("⟦MEDALHA⟧", _MEDALHA_SVG)
            .replace("⟦ESCOLA⟧", _esc_html(escola_nome))
            .replace("⟦ALUNO⟧", _esc_html(aluno_nome))
            .replace("⟦TURMA⟧", _esc_html(turma or "—"))
            .replace("⟦NOTA⟧", f"{nota_geral:.1f}".replace(".", ","))
            .replace("⟦POSICAO⟧", pos_txt)
            .replace("⟦ANO⟧", str(ano_letivo))
            .replace("⟦DATA⟧", datetime.now(timezone.utc).strftime("%d/%m/%Y")))
    doc = f"<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'></head><body>{html}</body></html>"

    with sync_playwright() as p:
        try:
            navegador = p.chromium.launch()
        except Exception:  # noqa: BLE001 — sem chromium baixado: tenta o Chrome do sistema
            navegador = p.chromium.launch(channel="chrome")
        try:
            pagina = navegador.new_page()
            pagina.set_content(doc, wait_until="networkidle")
            pdf = pagina.pdf(prefer_css_page_size=True, print_background=True,
                             landscape=True)
        finally:
            navegador.close()
    return pdf


def _certificado_fpdf(escola_nome: str, cor: str, aluno_nome: str,
                      turma: str, posicao: int | None, nota_geral: float,
                      ano_letivo: int) -> bytes:
    """Reserva simples (sem navegador): moldura + texto, 100% em Python."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    r, g, b = _hex_para_rgb(cor)

    pdf.set_draw_color(r, g, b)
    pdf.set_line_width(1.2)
    pdf.rect(8, 8, 281, 194)
    pdf.set_line_width(0.3)
    pdf.rect(12, 12, 273, 186)

    pdf.set_text_color(r, g, b)
    pdf.set_font("helvetica", "B", 30)
    pdf.set_y(40)
    pdf.cell(0, 14, "CERTIFICADO", align="C")

    pdf.set_text_color(60, 60, 60)
    pdf.set_font("helvetica", "", 13)
    pdf.set_y(62)
    pdf.cell(0, 8, _latin1(f"A escola {escola_nome} certifica que"), align="C")

    pdf.set_text_color(20, 20, 20)
    pdf.set_font("helvetica", "B", 26)
    pdf.set_y(78)
    # Nomes completos (lista de matrículas) podem ser longos: reduz a fonte
    # até caber na moldura, em vez de vazar pela borda.
    nome_cert = _latin1(aluno_nome)
    while pdf.get_string_width(nome_cert) > 250 and pdf.font_size_pt > 14:
        pdf.set_font_size(pdf.font_size_pt - 2)
    pdf.cell(0, 14, nome_cert, align="C")

    pdf.set_text_color(60, 60, 60)
    pdf.set_font("helvetica", "", 13)
    pdf.set_y(98)
    detalhe = _latin1(
        f"da turma {turma}, alcançou a nota geral {nota_geral:.1f}".replace(".", ","))
    if posicao:
        detalhe += f" e a {posicao}ª posição no Ranking Geral"
    pdf.cell(0, 8, detalhe, align="C")
    pdf.set_y(108)
    pdf.cell(0, 8, f"no Constela Edu - ano letivo de {ano_letivo}.", align="C")

    pdf.set_y(150)
    pdf.set_font("helvetica", "", 11)
    emitido = datetime.now(timezone.utc).strftime("%d/%m/%Y")
    pdf.cell(0, 8, f"Emitido em {emitido}", align="C")

    pdf.set_y(168)
    pdf.set_draw_color(120, 120, 120)
    pdf.line(110, 172, 187, 172)
    pdf.set_y(174)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, "Direção", align="C")

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Cartaz do Ranking Geral (pôster): documento "de vitrine" com a MESMA marca do
# certificado (moldura navy+dourado, 3 logos), a TABELA COMPLETA (todos os
# alunos, paginada automaticamente pelo Chromium) e um rodapé comemorativo.
# ---------------------------------------------------------------------------

def estatisticas_cartaz(db: Session, escola_id: int) -> dict:
    """Totais para os cartões do pôster: atividades (Matific) e livros
    (Elefante).

    Soma o SNAPSHOT ATUAL (o mais recente por data de referência — o mesmo
    critério usado no ranking e no painel, via scoring._snapshots_atuais), e só
    da MESMA população da tabela: alunos ATIVOS matriculados no ano letivo
    corrente. Assim os números batem com a tabela e com "estudantes ranqueados",
    sem inflar com alunos de anos anteriores/arquivados nem com correções para
    menos. Best-effort: qualquer falha vira zero — o cartaz nunca quebra por
    causa das estatísticas."""
    try:
        escola = db.get(Escola, escola_id)
        cohort = set(db.execute(
            select(Aluno.id)
            .join(Matricula, (Matricula.aluno_id == Aluno.id)
                  & (Matricula.ano_letivo == escola.ano_letivo_ativo))
            .where(Aluno.escola_id == escola_id, Aluno.status == "ativo")
        ).scalars().all())
        m_atuais = scoring._snapshots_atuais(db, escola_id, SnapshotMatific)
        e_atuais = scoring._snapshots_atuais(db, escola_id, SnapshotElefante)
        atividades = sum(s.atividades for aid, s in m_atuais.items() if aid in cohort)
        livros = sum(s.livros_unicos for aid, s in e_atuais.items() if aid in cohort)
        return {"atividades": int(atividades), "livros": int(livros)}
    except Exception:  # noqa: BLE001 — cartões decorativos: nunca derrubam o cartaz
        return {"atividades": 0, "livros": 0}


_CARTAZ_TEMPLATE = """<style>
  @page { size: A4 portrait; margin: 13mm 12mm 12mm; }
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family: Georgia,'Times New Roman',serif; color:#25304a;
         -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  .barra { height:5px; border-radius:3px;
           background:linear-gradient(90deg,#1B2A4A,#C9A24B 50%,#1B2A4A); }
  .cabecalho { display:flex; align-items:center; justify-content:space-between; gap:16px; margin:12px 2px 2px; }
  .cabecalho .logo-cidade { height:60px; width:60px; }
  .cabecalho .logo-pref { height:50px; width:auto; }
  .cabecalho .constela { display:flex; flex-direction:column; align-items:center; gap:2px; }
  .cabecalho .constela .nome { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-weight:800;
                 font-size:15px; letter-spacing:.5px; color:#1B2A4A; }
  .cabecalho .constela .nome b { color:#F5B942; }
  h1 { text-align:center; font-size:33px; font-weight:700; letter-spacing:7px; color:#1B2A4A; margin-top:8px; }
  .divisor { display:flex; align-items:center; justify-content:center; gap:10px; margin:6px 0 10px; }
  .divisor .linha { width:80px; height:1.5px; background:#C9A24B; }
  .divisor .losango { width:8px; height:8px; background:#F5B942; transform:rotate(45deg); }
  .pills { display:flex; justify-content:center; gap:10px; flex-wrap:wrap; margin-bottom:12px; }
  .pill { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-size:10.5px; letter-spacing:.6px;
          text-transform:uppercase; padding:5px 14px; border-radius:20px; }
  .pill.ano { background:#1B2A4A; color:#fff; }
  .pill.escola { background:#F3ECDA; color:#8A6D2B; border:1px solid #E4D3A6; }
  table { width:100%; border-collapse:collapse; font-size:11px; }
  thead { display:table-header-group; }
  thead th { background:#1B2A4A; color:#fff; font-family:'Trebuchet MS','Segoe UI',sans-serif;
             font-weight:700; font-size:9.5px; letter-spacing:.4px; text-transform:uppercase;
             padding:8px 6px; text-align:center; }
  thead th.esq { text-align:left; }
  .ic { vertical-align:middle; margin-right:4px; }
  tbody td { padding:5.5px 6px; border-bottom:1px solid #ECEEF2; text-align:center;
             font-variant-numeric:tabular-nums; }
  tbody td.aluno { text-align:left; font-weight:600; color:#25304a; }
  tbody td.turma, tbody td.serie { text-align:left; color:#5a6478; }
  tbody td.geral { font-weight:800; color:#1B2A4A; }
  tbody tr.top td { background:#FBF6E9; }
  tbody tr:nth-child(even):not(.top) td { background:#FAFBFC; }
  .medalha { display:inline-flex; width:23px; height:23px; border-radius:50%; align-items:center;
             justify-content:center; font-family:'Trebuchet MS',sans-serif; font-weight:800;
             font-size:12px; color:#1B2A4A; }
  .m1 { background:linear-gradient(#F8DA7B,#E3A62A); }
  .m2 { background:linear-gradient(#E5E9EF,#B7BFCC); }
  .m3 { background:linear-gradient(#E7BE86,#C07F3C); color:#fff; }
  .rodape { break-inside:avoid; }
  .tiles { display:flex; gap:10px; margin:16px 0 12px; }
  .tile { flex:1; text-align:center; border:1px solid #E7E3D4; border-radius:10px; padding:12px 6px; background:#FCFAF4; }
  .tile .n { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-size:23px; font-weight:800; color:#1B2A4A; }
  .tile .r { font-size:9px; letter-spacing:.5px; text-transform:uppercase; color:#8A6D2B; margin-top:3px; }
  .banner { display:flex; align-items:center; justify-content:center; gap:8px; text-align:center;
            background:#1B2A4A; color:#fff; border-radius:10px; padding:12px 16px; font-size:12px; letter-spacing:.3px; }
  .banner b { color:#F5D98A; }
  .frase { text-align:center; font-style:italic; color:#6b7488; font-size:12.5px; margin:13px 0 4px; }
  .frase::before { content:'“'; color:#C9A24B; } .frase::after { content:'”'; color:#C9A24B; }
  .rodape-marca { text-align:center; font-family:'Trebuchet MS','Segoe UI',sans-serif; font-size:9.5px;
                  letter-spacing:1.5px; text-transform:uppercase; color:#1B2A4A;
                  margin-top:10px; padding-top:9px; border-top:1px solid #E7E3D4; }
  .gerado { text-align:center; font-size:8.5px; color:#9aa2b1; margin-top:4px; }
</style>
<div class="barra"></div>
⟦CABECALHO⟧
<h1>RANKING GERAL</h1>
<div class="divisor"><span class="linha"></span><span class="losango"></span><span class="linha"></span></div>
<div class="pills">
  <span class="pill ano">Constela Edu · Ano letivo de ⟦ANO⟧</span>
  <span class="pill escola">⟦ESCOLA⟧</span>
</div>
<table>
  <thead><tr>
    <th>Posição</th><th class="esq">Aluno</th><th class="esq">Turma</th><th class="esq">Série</th>
    <th>⟦IC_STAR⟧Nota Matific</th><th>⟦IC_ELE⟧Nota Elefante</th><th>⟦IC_TRO⟧Nota Geral</th>
  </tr></thead>
  <tbody>⟦LINHAS⟧</tbody>
</table>
<div class="rodape">
  <div class="tiles">⟦TILES⟧</div>
  <div class="banner">⟦IC_STAR⟧<span><b>Parabéns</b> a todos os estudantes por cada passo na jornada do conhecimento!</span></div>
  <div class="frase">Cada conquista é uma estrela no caminho do conhecimento</div>
  <div class="rodape-marca">Constela Edu — Conectando aprendizagem, inspirando futuros</div>
  <div class="gerado">Gerado em ⟦DATA⟧</div>
</div>
"""


def _linhas_cartaz_html(linhas: list[list]) -> str:
    partes = []
    for linha in linhas:
        pos, nome, turma, serie, nm, ne, ng = (linha + [""] * 7)[:7]
        try:
            pos_int = int(pos)
        except (TypeError, ValueError):
            pos_int = 0
        if 1 <= pos_int <= 3:
            classe = {1: "m1", 2: "m2", 3: "m3"}[pos_int]
            pos_html = f'<span class="medalha {classe}">{pos_int}</span>'
            tr_classe = "top"
        else:
            pos_html = f"{pos_int}º"
            tr_classe = ""
        partes.append(
            f'<tr class="{tr_classe}"><td class="pos">{pos_html}</td>'
            f'<td class="aluno">{_esc_html(nome)}</td>'
            f'<td class="turma">{_esc_html(turma)}</td>'
            f'<td class="serie">{_esc_html(serie)}</td>'
            f'<td>{_fmt_nota(nm)}</td><td>{_fmt_nota(ne)}</td>'
            f'<td class="geral">{_fmt_nota(ng)}</td></tr>')
    return "".join(partes)


def _faixa_anos(linhas: list[list]) -> tuple[str, str]:
    """Deriva 'N' e rótulo do 4º cartão a partir das séries dos alunos
    ('2º–5º' / 'anos atendidos'). Sem séries numéricas, cai para nº de turmas."""
    anos = []
    for linha in linhas:
        m = re.match(r"\s*(\d+)", str(linha[3]) if len(linha) > 3 else "")
        if m:
            anos.append(int(m.group(1)))
    if anos:
        lo, hi = min(anos), max(anos)
        return (f"{lo}º–{hi}º", "anos atendidos") if lo != hi else (f"{lo}º", "ano")
    turmas = len({linha[2] for linha in linhas if len(linha) > 2})
    return str(turmas), ("turmas" if turmas != 1 else "turma")


def gerar_cartaz_ranking(escola_nome: str, cor: str, ano_letivo: int,
                         cabecalho: list[str], linhas: list[list],
                         estatisticas: dict,
                         logos: tuple[str, str] | None = None) -> bytes:
    """Cartaz do Ranking Geral em PDF (todos os alunos). HTML→Chromium com o
    layout completo; sem navegador, cai para o PDF simples de ranking (reserva)."""
    try:
        return _cartaz_html_pdf(escola_nome, ano_letivo, linhas, estatisticas, logos)
    except Exception:  # noqa: BLE001 — sem Chromium/erro: PDF simples de ranking
        return gerar_pdf("Ranking Geral", escola_nome, cor, cabecalho, linhas)


def _cartaz_html_pdf(escola_nome: str, ano_letivo: int, linhas: list[list],
                     estatisticas: dict, logos: tuple[str, str] | None = None) -> bytes:
    from playwright.sync_api import sync_playwright

    def _mil(n) -> str:
        return f"{int(n):,}".replace(",", ".")

    faixa_n, faixa_r = _faixa_anos(linhas)
    tiles = [
        (_mil(estatisticas.get("atividades", 0)), "atividades"),
        (_mil(estatisticas.get("livros", 0)), "livros lidos"),
        (_mil(len(linhas)), "estudantes ranqueados"),
        (faixa_n, faixa_r),
    ]
    tiles_html = "".join(
        f'<div class="tile"><div class="n">{n}</div><div class="r">{r}</div></div>'
        for n, r in tiles)

    html = (_CARTAZ_TEMPLATE
            .replace("⟦CABECALHO⟧", _cabecalho_logos(logos))
            .replace("⟦IC_STAR⟧", _IC_STAR)
            .replace("⟦IC_ELE⟧", _IC_ELEFANTE)
            .replace("⟦IC_TRO⟧", _IC_TROFEU)
            .replace("⟦ANO⟧", str(ano_letivo))
            .replace("⟦ESCOLA⟧", _esc_html(escola_nome))
            .replace("⟦LINHAS⟧", _linhas_cartaz_html(linhas))
            .replace("⟦TILES⟧", tiles_html)
            .replace("⟦DATA⟧", datetime.now(timezone.utc).strftime("%d/%m/%Y")))
    doc = (f"<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'>"
           f"</head><body>{html}</body></html>")

    with sync_playwright() as p:
        try:
            navegador = p.chromium.launch()
        except Exception:  # noqa: BLE001 — sem chromium baixado: tenta o Chrome do sistema
            navegador = p.chromium.launch(channel="chrome")
        try:
            pagina = navegador.new_page()
            pagina.set_content(doc, wait_until="networkidle")
            pdf = pagina.pdf(prefer_css_page_size=True, print_background=True)
        finally:
            navegador.close()
    return pdf


# ---------------------------------------------------------------------------
# Relatórios "de vitrine" em PDF — Lista de Alunos e Catálogo de Livros.
# Mesma família visual dos demais documentos (faixa com os 3 logos, título,
# cartões de estatística, tabelas navy+zebra), com rodapé/numeração de página
# em TODAS as páginas via header/footer do Chromium. Reserva: gerar_pdf simples.
# ---------------------------------------------------------------------------

_REPORT_CSS = """<style>
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family: Georgia,'Times New Roman',serif; color:#25304a;
         -webkit-print-color-adjust:exact; print-color-adjust:exact; }
  .faixa { border:1.5px solid #1B2A4A; border-left:6px solid #C9A24B; border-radius:6px;
           padding:9px 22px; }
  .faixa .cabecalho { display:flex; align-items:center; justify-content:space-between; width:100%; gap:16px; }
  .faixa .logo-cidade { height:48px; width:48px; }
  .faixa .logo-pref { height:34px; width:auto; }
  .faixa .constela { display:flex; flex-direction:column; align-items:center; gap:1px; }
  .faixa .constela .nome { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-weight:800;
                 font-size:19px; letter-spacing:1px; color:#1B2A4A; }
  .faixa .constela .nome b { color:#F5B942; }
  h1 { text-align:center; font-size:30px; font-weight:700; letter-spacing:3px; color:#1B2A4A; margin-top:14px; }
  .escola-nome { text-align:center; font-size:13px; font-weight:700; letter-spacing:.4px; color:#1B2A4A; margin-top:7px; }
  .subtitulo { text-align:center; font-size:11.5px; color:#6b7488; margin-top:3px; }
  .tiles { display:flex; border:1px solid #E4D3A6; border-radius:8px; overflow:hidden; margin:14px 0 10px; }
  .tile { flex:1; padding:9px 16px; border-left:1px solid #EFE9DA; }
  .tile:first-child { border-left:0; }
  .tile .n { font-family:'Trebuchet MS','Segoe UI',sans-serif; font-size:17px; font-weight:800; color:#1B2A4A; }
  .tile .r { font-size:8.5px; letter-spacing:.4px; text-transform:uppercase; color:#8A6D2B; margin-top:2px; }
  .intro { font-size:10.5px; color:#6b7488; line-height:1.5; margin-bottom:4px; }
  .secao { font-size:15px; font-weight:700; color:#1B2A4A; letter-spacing:1px; margin:16px 0 6px; break-after:avoid; }
  table.tab { width:100%; border-collapse:collapse; font-size:10.5px; }
  table.tab thead { display:table-header-group; }
  table.tab thead th { background:#1B2A4A; color:#fff; font-family:'Trebuchet MS','Segoe UI',sans-serif;
             font-weight:700; font-size:8.5px; letter-spacing:.4px; text-transform:uppercase;
             padding:8px 10px; text-align:center; }
  table.tab thead th.esq { text-align:left; }
  table.tab tbody td { padding:6px 10px; border-bottom:1px solid #ECEEF2; text-align:center;
             font-variant-numeric:tabular-nums; }
  table.tab tbody td.esq { text-align:left; }
  table.tab tbody td.num { color:#8a93a5; width:38px; }
  table.tab tbody td.nome { font-weight:600; color:#25304a; }
  table.tab tbody tr:nth-child(even) td { background:#F7F8FA; }
  .sit { font-weight:700; color:#5a6478; }
  .sit.ativo { color:#2FA24B; }
  .sit .dot { display:inline-block; width:7px; height:7px; border-radius:50%;
              background:#2FA24B; margin-right:5px; vertical-align:middle; }
</style>"""

# Rodapé com numeração — renderizado pelo Chromium em TODAS as páginas. Precisa
# de estilo inline (o CSS da página não se aplica ao template de rodapé).
_REPORT_FOOTER = (
    '<div style="width:100%; box-sizing:border-box; padding:4px 11mm 0; margin:0 6mm;'
    ' border-top:1px solid #C9A24B; font-family:Georgia,serif; font-size:8px;'
    ' color:#8a93a5; display:flex; justify-content:space-between; align-items:center;">'
    '<span>CONSTELA EDU &bull; Conectando aprendizagem, inspirando futuros.</span>'
    '<span>Página <span class="pageNumber"></span></span></div>')


def _report_tiles(tiles: list[tuple[str, str]]) -> str:
    return ('<div class="tiles">' + "".join(
        f'<div class="tile"><div class="n">{_esc_html(n)}</div>'
        f'<div class="r">{_esc_html(r)}</div></div>' for n, r in tiles) + '</div>')


def _relatorio_shell(titulo: str, escola_nome: str, subtitulo: str,
                     tiles: list[tuple[str, str]], intro: str, corpo: str,
                     logos: tuple[str, str] | None = None) -> str:
    return (_REPORT_CSS
            + '<div class="faixa">' + _cabecalho_logos(logos) + '</div>'
            + f'<h1>{_esc_html(titulo)}</h1>'
            + f'<p class="escola-nome">{_esc_html(escola_nome)}</p>'
            + f'<p class="subtitulo">{subtitulo}</p>'
            + _report_tiles(tiles)
            + f'<p class="intro">{intro}</p>'
            + corpo)


def _relatorio_html_pdf(html_body: str) -> bytes:
    """Renderiza um relatório 'de vitrine' (A4 retrato) com rodapé/numeração."""
    from playwright.sync_api import sync_playwright

    doc = (f"<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'>"
           f"</head><body>{html_body}</body></html>")
    with sync_playwright() as p:
        try:
            navegador = p.chromium.launch()
        except Exception:  # noqa: BLE001 — sem chromium baixado: tenta o Chrome do sistema
            navegador = p.chromium.launch(channel="chrome")
        try:
            pagina = navegador.new_page()
            pagina.set_content(doc, wait_until="networkidle")
            pdf = pagina.pdf(
                format="A4", print_background=True, display_header_footer=True,
                header_template="<div></div>", footer_template=_REPORT_FOOTER,
                margin={"top": "12mm", "bottom": "16mm", "left": "11mm", "right": "11mm"})
        finally:
            navegador.close()
    return pdf


def _faixa_anos_relatorio(series) -> tuple[str, str]:
    anos = [int(m.group(1)) for s in series
            if (m := re.match(r"\s*(\d+)", str(s)))]
    if anos:
        lo, hi = min(anos), max(anos)
        return (f"{lo}º → {hi}º", "anos") if lo != hi else (f"{lo}º", "ano")
    return str(len({s for s in series if s})), "séries"


def gerar_lista_alunos_pdf(escola_nome: str, cor: str, ano_letivo: int,
                           cabecalho: list[str], linhas: list[list],
                           logos: tuple[str, str] | None = None) -> bytes:
    """Lista de Alunos em PDF de vitrine (agrupada por turma). Reserva: PDF simples."""
    try:
        return _relatorio_html_pdf(_shell_lista_alunos(escola_nome, ano_letivo, linhas, logos))
    except Exception:  # noqa: BLE001 — sem Chromium/erro: PDF simples
        return gerar_pdf("Lista de Alunos", escola_nome, cor, cabecalho, linhas)


def _shell_lista_alunos(escola_nome: str, ano_letivo: int, linhas: list[list],
                        logos: tuple[str, str] | None = None) -> str:
    from itertools import groupby

    # linhas de linhas_alunos: [Aluno, Turma, Série, Nº de chamada, Situação]
    total = len(linhas)
    turmas = {linha[1] for linha in linhas}
    faixa_n, faixa_r = _faixa_anos_relatorio([linha[2] for linha in linhas])
    ativos = sum(1 for linha in linhas if str(linha[4]).lower().startswith("ativ"))
    pct = f"{round(100 * ativos / total)}%" if total else "—"
    tiles = [(str(total), "alunos cadastrados"), (str(len(turmas)), "turmas"),
             (faixa_n, faixa_r), (pct, "status ativo")]

    corpo = []
    for turma, grupo in groupby(linhas, key=lambda linha: linha[1]):
        corpo.append(f'<h2 class="secao">{_esc_html(turma)}</h2>')
        corpo.append('<table class="tab"><thead><tr>'
                     '<th>Nº</th><th class="esq">Nome do estudante</th>'
                     '<th>Série</th><th>Nº de chamada</th><th>Situação</th>'
                     '</tr></thead><tbody>')
        for i, linha in enumerate(grupo, 1):
            nome, _turma, serie, chamada, status = (list(linha) + [""] * 5)[:5]
            ativo = str(status).lower().startswith("ativ")
            chamada_txt = str(chamada) if chamada not in (None, "", 0, "0") else "—"
            dot = '<span class="dot"></span>' if ativo else ''
            corpo.append(
                f'<tr><td class="num">{i:02d}</td>'
                f'<td class="esq nome">{_esc_html(nome)}</td>'
                f'<td>{_esc_html(serie)}</td>'
                f'<td>{_esc_html(chamada_txt)}</td>'
                f'<td class="sit {"ativo" if ativo else ""}">{dot}'
                f'{_esc_html(str(status).capitalize())}</td></tr>')
        corpo.append('</tbody></table>')

    subtitulo = (f"Cadastro escolar · Ano letivo de {ano_letivo} · "
                 f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y')}")
    intro = ("Relação oficial de estudantes cadastrados no Constela Edu. A situação "
             "corresponde ao status registrado no momento da emissão deste relatório.")
    return _relatorio_shell("Lista de Alunos", escola_nome, subtitulo, tiles,
                            intro, "".join(corpo), logos)


def gerar_catalogo_livros_pdf(escola_nome: str, cor: str, ano_letivo: int,
                              cabecalho: list[str], linhas: list[list],
                              logos: tuple[str, str] | None = None) -> bytes:
    """Catálogo de Livros em PDF de vitrine. Reserva: PDF simples."""
    try:
        return _relatorio_html_pdf(_shell_catalogo(escola_nome, ano_letivo, linhas, logos))
    except Exception:  # noqa: BLE001 — sem Chromium/erro: PDF simples
        return gerar_pdf("Catálogo de Livros", escola_nome, cor, cabecalho, linhas)


def _shell_catalogo(escola_nome: str, ano_letivo: int, linhas: list[list],
                    logos: tuple[str, str] | None = None) -> str:
    # linhas de linhas_livros: [Título, Autor, Nível, Categoria, Leituras]
    total = len(linhas)
    autores = {linha[1] for linha in linhas if len(linha) > 1 and linha[1]}
    niveis = {linha[2] for linha in linhas if len(linha) > 2 and linha[2]}
    leituras = sum(int(linha[4] or 0) for linha in linhas if len(linha) > 4)
    tiles = [(str(total), "livros no acervo"), (str(len(autores)), "autores"),
             (str(len(niveis)), "níveis"),
             (f"{leituras:,}".replace(",", "."), "leituras")]

    corpo = ['<table class="tab"><thead><tr>'
             '<th class="esq">Título</th><th class="esq">Autor</th><th>Nível</th>'
             '<th class="esq">Categoria</th><th>Leituras</th>'
             '</tr></thead><tbody>']
    for linha in linhas:
        titulo, autor, nivel, categoria, leituras_n = (list(linha) + [""] * 5)[:5]
        corpo.append(
            f'<tr><td class="esq nome">{_esc_html(titulo)}</td>'
            f'<td class="esq">{_esc_html(autor or "—")}</td>'
            f'<td>{_esc_html(nivel or "—")}</td>'
            f'<td class="esq">{_esc_html(categoria or "—")}</td>'
            f'<td>{_esc_html(leituras_n)}</td></tr>')
    corpo.append('</tbody></table>')

    subtitulo = (f"Acervo de leitura · Ano letivo de {ano_letivo} · "
                 f"Gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y')}")
    intro = ("Catálogo de livros disponíveis no Constela Edu, com o total de "
             "leituras registradas por título.")
    return _relatorio_shell("Catálogo de Livros", escola_nome, subtitulo, tiles,
                            intro, "".join(corpo), logos)
